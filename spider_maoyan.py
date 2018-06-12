import requests
import bs4
from requests.exceptions import RequestException
import openpyxl


def get_one_page(url, headers):
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.text
        return None
    except RequestException:
        return None


def parse_one_page(html):
    soup = bs4.BeautifulSoup(html, 'lxml')
    # 获取电影名
    movies = []
    targets = soup.find_all(class_='name')
    for each in targets:
        movies.append(each.get_text())
    # 获取评分
    scores = []
    targets = soup.find_all(class_='score')
    for each in targets:
        scores.append(each.get_text())
    # 获取主演信息
    star_message = []
    targets = soup.find_all(class_='star')
    for each in targets:
        star_message.append(each.get_text().split('\n')[1].strip())
        print(each.get_text().split('\n')[1].strip())
    # 获取上映时间
    play_time = []
    targets = soup.find_all(class_='releasetime')
    for each in targets:
        play_time.append(each.get_text())
    result = []
    length = len(movies)
    for j in range(length):
        result.append([movies[j], scores[j], star_message[j], play_time[j]])

    return result


def save_to_excel(result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '电影名称'
    ws['B1'] = '评分'
    ws['C1'] = '主演'
    ws['D1'] = '上映时间'
    for item in result:
        ws.append(item)
    wb.save('猫眼电影TOP100.xlsx')


def main():
    result = []
    for i in range(10):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.116 Safari/537.36',
        }
        url = 'http://maoyan.com/board/4?offset=' + str(i * 10)
        html = get_one_page(url, headers)
        result.extend(parse_one_page(html))
    save_to_excel(result)


if __name__ == '__main__':
    # for i in range(10):
    #     main(offset=i * 10)
    #     time.sleep(1)
    main()
